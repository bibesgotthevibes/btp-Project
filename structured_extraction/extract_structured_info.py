"""
extract_structured_info.py
──────────────────────────
Main extraction script that reads discharge summaries from
anotated_dataset_v2.xlsx and extracts structured medical information
using Groq and Cerebras LLM APIs.

Usage:
    python extract_structured_info.py
"""

import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
import io

# ── Fix Windows console encoding for emoji/unicode output ────────────────────
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── Ensure imports work when run from this directory ──────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from prompt_template import build_extraction_messages
from schema import validate_extraction, extraction_summary

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
CEREBRAS_API_KEY = os.getenv("CEREBRAS_API_KEY", "")

GROQ_ENDPOINT = "https://api.groq.com/openai/v1/chat/completions"
CEREBRAS_ENDPOINT = "https://api.cerebras.ai/v1/chat/completions"

# Primary: Groq with Qwen 3.6 27B
GROQ_MODEL = "qwen/qwen3.6-27b"
# Secondary: Groq with GPT OSS 20B (different model for comparison)
GROQ_MODEL_2 = "openai/gpt-oss-20b"
# Cerebras model (requires billing — used as fallback)
CEREBRAS_MODEL = "gemma-4-31b"

# Max retries for rate limit (429) errors
MAX_RETRIES = 3
RETRY_BASE_DELAY = 15  # seconds

# How many summaries to process (from the 200 in the dataset)
NUM_SUMMARIES = 5

# Which row indices to pick (0-indexed, after header).
# We pick a diverse set: short stay, long complex, death case, etc.
SAMPLE_INDICES = [0, 1, 2, 3, 9]  # rows 1,2,3,4,10 from the dataset

RESULTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "anotated_dataset_v2.xlsx")

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL READER (using openpyxl — no pandas dependency required)
# ══════════════════════════════════════════════════════════════════════════════

def read_discharge_summaries(excel_path: str, indices: list[int]) -> list[dict]:
    """
    Read specific rows from the Excel file and return a list of dicts
    with the discharge summary text, diagnosis ICD, outcome, etc.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # Read headers from row 1
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    print(f"📋 Found columns: {headers}")
    print(f"📊 Sheet has {ws.max_row - 1} data rows\n")

    # Find column indices by partial match
    summary_col = None
    diagnosis_col = None
    outcome_col = None
    stay_col = None
    specialty_col = None

    for i, h in enumerate(headers):
        hl = h.lower()
        if "discharge" in hl and "summary" in hl:
            summary_col = i
        elif "diagnosis" in hl or "icd" in hl:
            diagnosis_col = i
        elif "outcome" in hl:
            outcome_col = i
        elif "stay" in hl or "length" in hl or "lenght" in hl:
            stay_col = i
        elif "special" in hl:
            specialty_col = i

    if summary_col is None:
        print("ERROR: Could not find discharge summary column!")
        print(f"  Available headers: {headers}")
        sys.exit(1)

    print(f"✅ Using column {summary_col} ('{headers[summary_col]}') for discharge text")

    # Read all data rows into memory (since read_only mode uses generator)
    all_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        all_rows.append(row)

    wb.close()

    # Pick the requested indices
    results = []
    for idx in indices:
        if idx >= len(all_rows):
            print(f"⚠️  Skipping index {idx} — only {len(all_rows)} rows available")
            continue

        row = all_rows[idx]
        summary_text = str(row[summary_col]) if row[summary_col] else ""
        diagnosis_text = str(row[diagnosis_col]) if diagnosis_col and row[diagnosis_col] else ""
        outcome_text = str(row[outcome_col]) if outcome_col and row[outcome_col] else ""
        stay_text = str(row[stay_col]) if stay_col and row[stay_col] else ""

        if not summary_text.strip() or summary_text == "None":
            print(f"⚠️  Skipping index {idx} — empty summary")
            continue

        results.append({
            "row_index": idx + 2,  # 1-indexed Excel row (header is row 1)
            "summary_text": summary_text.strip(),
            "diagnosis_icd": diagnosis_text.strip(),
            "outcome": outcome_text.strip(),
            "stay": stay_text.strip(),
        })

    return results


# ══════════════════════════════════════════════════════════════════════════════
# LLM API CALLERS
# ══════════════════════════════════════════════════════════════════════════════

def call_llm_api(
    endpoint: str,
    api_key: str,
    model: str,
    messages: list[dict],
    temperature: float = 0.1,
    max_tokens: int = 4096,
) -> tuple[str, float]:
    """
    Call an OpenAI-compatible chat completions endpoint.
    Uses 'requests' library if available (better TLS/header handling),
    falls back to urllib.
    Returns (response_text, elapsed_seconds).
    """
    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    # For Qwen models, suppress reasoning/thinking output
    if "qwen" in model.lower():
        payload["reasoning_format"] = "hidden"

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
        "User-Agent": "MedSimplify-Extraction/1.0",
    }

    start = time.time()

    # ── Try with 'requests' library first ─────────────────────────────────
    try:
        import requests as req_lib
        resp = req_lib.post(
            endpoint,
            json=payload,
            headers=headers,
            timeout=120,
        )
        elapsed = time.time() - start

        if resp.status_code != 200:
            raise RuntimeError(
                f"API Error {resp.status_code}: {resp.text[:500]}"
            )

        body = resp.json()
        text = body.get("choices", [{}])[0].get("message", {}).get("content", "")
        return text, elapsed

    except ImportError:
        pass  # Fall through to urllib

    # ── Fallback to urllib ────────────────────────────────────────────────
    payload_bytes = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        endpoint,
        data=payload_bytes,
        headers=headers,
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8") if e.fp else ""
        raise RuntimeError(
            f"API Error {e.code}: {error_body[:500]}"
        ) from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"Connection Error: {e.reason}") from e

    elapsed = time.time() - start
    text = body.get("choices", [{}])[0].get("message", {}).get("content", "")
    return text, elapsed


def parse_json_response(raw_text: str) -> dict | None:
    """
    Parse JSON from LLM response. Handles cases where the LLM wraps
    JSON in markdown code fences, <think> reasoning blocks, or adds extra text.
    """
    # Strip <think>...</think> reasoning blocks (Qwen 3.6 etc.)
    cleaned = re.sub(r"<think>[\s\S]*?</think>", "", raw_text).strip()
    # Also strip if there's an unclosed <think> block at the start
    cleaned = re.sub(r"^<think>[\s\S]*", "", cleaned).strip()

    # Try direct parse first (on cleaned text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # Try extracting from markdown code fences
    patterns = [
        r"```json\s*\n([\s\S]*?)\n\s*```",
        r"```\s*\n([\s\S]*?)\n\s*```",
        r"(\{[\s\S]*\})",
    ]
    for pattern in patterns:
        match = re.search(pattern, cleaned)
        if match:
            try:
                candidate = match.group(1) if match.lastindex else match.group(0)
                return json.loads(candidate)
            except json.JSONDecodeError:
                continue

    # Last resort: try on the original raw text too
    for pattern in patterns:
        match = re.search(pattern, raw_text)
        if match:
            try:
                candidate = match.group(1) if match.lastindex else match.group(0)
                return json.loads(candidate)
            except json.JSONDecodeError:
                continue

    return None


# ══════════════════════════════════════════════════════════════════════════════
# MAIN EXTRACTION LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def extract_with_provider(
    provider_name: str,
    endpoint: str,
    api_key: str,
    model: str,
    summaries: list[dict],
) -> dict:
    """
    Run extraction for all summaries using a single provider.
    Returns the full results dict.
    """
    results = {
        "provider": provider_name,
        "model": model,
        "extraction_timestamp": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "num_summaries_processed": 0,
        "num_successful": 0,
        "num_failed": 0,
        "total_time_seconds": 0,
        "extractions": [],
    }

    for i, summary_data in enumerate(summaries):
        summary_id = summary_data["row_index"]
        summary_text = summary_data["summary_text"]
        print(f"  [{i+1}/{len(summaries)}] Processing row {summary_id} "
              f"({len(summary_text)} chars)...", end=" ", flush=True)

        messages = build_extraction_messages(summary_text)

        entry = {
            "summary_id": summary_id,
            "original_summary_snippet": summary_text[:200] + "..." if len(summary_text) > 200 else summary_text,
            "dataset_diagnosis_icd": summary_data["diagnosis_icd"],
            "dataset_outcome": summary_data["outcome"],
            "dataset_stay": summary_data["stay"],
            "extraction": None,
            "validation": None,
            "extraction_summary": None,
            "error": None,
            "time_seconds": 0,
        }

        # Retry logic for rate limits (429)
        last_error = None
        for attempt in range(MAX_RETRIES + 1):
            try:
                raw_response, elapsed = call_llm_api(
                    endpoint=endpoint,
                    api_key=api_key,
                    model=model,
                    messages=messages,
                )
                entry["time_seconds"] = round(elapsed, 2)
                results["total_time_seconds"] += elapsed

                parsed = parse_json_response(raw_response)
                if parsed is None:
                    entry["error"] = f"Failed to parse JSON from response. Raw (first 500 chars): {raw_response[:500]}"
                    results["num_failed"] += 1
                    print(f"[!] JSON parse failed ({elapsed:.1f}s)")
                else:
                    is_valid, issues = validate_extraction(parsed)
                    entry["extraction"] = parsed
                    entry["validation"] = {
                        "is_valid": is_valid,
                        "issues": issues,
                    }
                    entry["extraction_summary"] = extraction_summary(parsed)
                    results["num_successful"] += 1
                    diag = parsed.get("diagnoses", {}).get("primary", "N/A")
                    n_meds = len(parsed.get("medications", []))
                    n_anom = len(parsed.get("anomalies_detected", []))
                    status = "[OK]" if is_valid else "[WARN]"
                    print(f"{status} Done ({elapsed:.1f}s) -- Diagnosis: {diag[:50]}, "
                          f"{n_meds} meds, {n_anom} anomalies")
                last_error = None
                break  # Success, exit retry loop

            except RuntimeError as e:
                last_error = e
                err_str = str(e)
                # Retry on rate limit (429) or request too large (413)
                if "429" in err_str and attempt < MAX_RETRIES:
                    wait_time = RETRY_BASE_DELAY * (attempt + 1)
                    print(f"[RATE LIMIT] Waiting {wait_time}s before retry {attempt+2}/{MAX_RETRIES+1}...",
                          flush=True)
                    time.sleep(wait_time)
                    continue
                elif "413" in err_str:
                    # Request too large — skip, no point retrying
                    entry["error"] = f"Request too large for model: {err_str[:200]}"
                    results["num_failed"] += 1
                    print(f"[!] Request too large, skipping")
                    break
                else:
                    entry["error"] = str(e)
                    results["num_failed"] += 1
                    print(f"[!] Error: {str(e)[:100]}")
                    break

            except Exception as e:
                entry["error"] = str(e)
                results["num_failed"] += 1
                print(f"[!] Error: {str(e)[:100]}")
                break

        if last_error and entry["error"] is None:
            entry["error"] = str(last_error)
            results["num_failed"] += 1

        results["extractions"].append(entry)
        results["num_summaries_processed"] += 1

        # Delay between calls to respect rate limits
        if i < len(summaries) - 1:
            time.sleep(3)

    results["total_time_seconds"] = round(results["total_time_seconds"], 2)
    return results


def build_comparison(groq_results: dict, cerebras_results: dict) -> dict:
    """Build a side-by-side comparison summary."""
    comparison = {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "providers_compared": [
            {
                "name": "Groq",
                "model": groq_results["model"],
                "successful": groq_results["num_successful"],
                "failed": groq_results["num_failed"],
                "total_time_seconds": groq_results["total_time_seconds"],
            },
            {
                "name": "Cerebras",
                "model": cerebras_results["model"],
                "successful": cerebras_results["num_successful"],
                "failed": cerebras_results["num_failed"],
                "total_time_seconds": cerebras_results["total_time_seconds"],
            },
        ],
        "per_summary_comparison": [],
    }

    for g_entry, c_entry in zip(groq_results["extractions"], cerebras_results["extractions"]):
        row = {
            "summary_id": g_entry["summary_id"],
            "groq": {
                "success": g_entry["error"] is None,
                "time_seconds": g_entry["time_seconds"],
                "summary": g_entry.get("extraction_summary"),
            },
            "cerebras": {
                "success": c_entry["error"] is None,
                "time_seconds": c_entry["time_seconds"],
                "summary": c_entry.get("extraction_summary"),
            },
        }
        comparison["per_summary_comparison"].append(row)

    return comparison


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("  STRUCTURED MEDICAL INFORMATION EXTRACTION")
    print("  Model 1: Groq (qwen/qwen3.6-27b)")
    print("  Model 2: Groq (openai/gpt-oss-20b)")
    print("=" * 70)
    print()

    # Resolve paths
    excel_path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at: {excel_path}")
        sys.exit(1)

    os.makedirs(RESULTS_DIR, exist_ok=True)

    # ── Step 1: Read discharge summaries ──────────────────────────────────────
    print("📖 Reading discharge summaries from Excel...")
    summaries = read_discharge_summaries(excel_path, SAMPLE_INDICES)
    print(f"📋 Loaded {len(summaries)} summaries for extraction\n")

    if not summaries:
        print("ERROR: No valid summaries found!")
        sys.exit(1)

    # ── Step 2: Extract with Groq Model 1 (Qwen 3.6 27B) ────────────────────
    print("-" * 70)
    print(f">> GROQ EXTRACTION - Model 1: {GROQ_MODEL}")
    print("-" * 70)
    groq_results = extract_with_provider(
        provider_name="Groq",
        endpoint=GROQ_ENDPOINT,
        api_key=GROQ_API_KEY,
        model=GROQ_MODEL,
        summaries=summaries,
    )
    groq_path = os.path.join(RESULTS_DIR, "groq_extractions.json")
    with open(groq_path, "w", encoding="utf-8") as f:
        json.dump(groq_results, f, indent=2, ensure_ascii=False)
    print(f"\n>> Groq Model 1 results saved to: {groq_path}")
    print(f"   OK: {groq_results['num_successful']} successful, "
          f"FAIL: {groq_results['num_failed']} failed, "
          f"Time: {groq_results['total_time_seconds']:.1f}s total\n")

    # Brief pause between models
    print(">> Waiting 10 seconds before Model 2 calls...\n")
    time.sleep(10)

    # ── Step 3: Extract with Groq Model 2 (GPT OSS 20B) ──────────────────────
    print("-" * 70)
    print(f">> GROQ EXTRACTION - Model 2: {GROQ_MODEL_2}")
    print("-" * 70)
    model2_results = extract_with_provider(
        provider_name="Groq-GPT-OSS",
        endpoint=GROQ_ENDPOINT,
        api_key=GROQ_API_KEY,
        model=GROQ_MODEL_2,
        summaries=summaries,
    )
    model2_path = os.path.join(RESULTS_DIR, "groq_gpt_oss_extractions.json")
    with open(model2_path, "w", encoding="utf-8") as f:
        json.dump(model2_results, f, indent=2, ensure_ascii=False)
    print(f"\n>> Groq Model 2 results saved to: {model2_path}")
    print(f"   OK: {model2_results['num_successful']} successful, "
          f"FAIL: {model2_results['num_failed']} failed, "
          f"Time: {model2_results['total_time_seconds']:.1f}s total\n")

    # ── Also try Cerebras if key is available ──────────────────────────────────
    cerebras_results = None
    try:
        print("-" * 70)
        print(f">> CEREBRAS EXTRACTION - Model: {CEREBRAS_MODEL}")
        print("-" * 70)
        cerebras_results = extract_with_provider(
            provider_name="Cerebras",
            endpoint=CEREBRAS_ENDPOINT,
            api_key=CEREBRAS_API_KEY,
            model=CEREBRAS_MODEL,
            summaries=summaries,
        )
        cerebras_path = os.path.join(RESULTS_DIR, "cerebras_extractions.json")
        with open(cerebras_path, "w", encoding="utf-8") as f:
            json.dump(cerebras_results, f, indent=2, ensure_ascii=False)
        print(f"\n>> Cerebras results saved to: {cerebras_path}")
        print(f"   OK: {cerebras_results['num_successful']} successful, "
              f"FAIL: {cerebras_results['num_failed']} failed, "
              f"Time: {cerebras_results['total_time_seconds']:.1f}s total\n")
    except Exception as e:
        print(f"\n>> Cerebras extraction skipped due to error: {e}\n")

    # ── Step 4: Build comparison ──────────────────────────────────────────────
    print("-" * 70)
    print(">> BUILDING COMPARISON SUMMARY")
    print("-" * 70)
    comparison = build_comparison(groq_results, model2_results)
    comparison_path = os.path.join(RESULTS_DIR, "comparison_summary.json")
    with open(comparison_path, "w", encoding="utf-8") as f:
        json.dump(comparison, f, indent=2, ensure_ascii=False)
    print(f">> Comparison saved to: {comparison_path}\n")

    # ── Final Summary ─────────────────────────────────────────────────────────
    print("=" * 70)
    print("  EXTRACTION COMPLETE")
    print("=" * 70)
    print(f"\n  Results saved in: {RESULTS_DIR}/")
    print(f"     - groq_extractions.json           ({groq_results['num_successful']}/{len(summaries)} successful)")
    print(f"     - groq_gpt_oss_extractions.json   ({model2_results['num_successful']}/{len(summaries)} successful)")
    if cerebras_results:
        print(f"     - cerebras_extractions.json       ({cerebras_results['num_successful']}/{len(summaries)} successful)")
    print(f"     - comparison_summary.json")
    print(f"\n  Groq Qwen time:    {groq_results['total_time_seconds']:.1f}s")
    print(f"  Groq GPT-OSS time: {model2_results['total_time_seconds']:.1f}s")
    if cerebras_results:
        print(f"  Cerebras time:     {cerebras_results['total_time_seconds']:.1f}s")
    print()


if __name__ == "__main__":
    main()
