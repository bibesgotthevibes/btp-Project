"""
make_report.py
==============
Run this script LOCALLY (no GPU / Kaggle needed) to turn the pipeline
output Excel into two easy-to-read files:

  1. results_report.html  — open in any browser (Chrome, Edge, Firefox)
  2. results_formatted.xlsx — Excel with wrapped text, wide columns, colours

Supports both:
  - English-only output  (lay_english_pipeline_output.xlsx)
  - Multilingual output  (multilingual_pipeline_output.xlsx) — adds language tabs

Usage:
    python make_report.py
    python make_report.py  results_2/lay_english_pipeline_output.xlsx
    python make_report.py  results_2/multilingual_pipeline_output.xlsx
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# ── Locate input file ─────────────────────────────────────────────
if len(sys.argv) > 1:
    INPUT_PATH = sys.argv[1]
else:
    # Auto-find the most recent output in known locations
    candidates = [
        r"results_2\multilingual_pipeline_output.xlsx",
        r"results_2\lay_english_pipeline_output.xlsx",
        r"results_1\lay_english_pipeline_output.xlsx",
        "multilingual_pipeline_output.xlsx",
        "lay_english_pipeline_output.xlsx",
    ]
    INPUT_PATH = next((p for p in candidates if os.path.exists(p)), None)
    if INPUT_PATH is None:
        print("ERROR: Could not find lay_english_pipeline_output.xlsx")
        print("Pass the path as an argument: python make_report.py <path>")
        sys.exit(1)

OUT_DIR  = os.path.dirname(os.path.abspath(INPUT_PATH))
HTML_OUT = os.path.join(OUT_DIR, "results_report.html")
XLSX_OUT = os.path.join(OUT_DIR, "results_formatted.xlsx")

print(f"Reading: {INPUT_PATH}")
df = pd.read_excel(INPUT_PATH)
print(f"  {len(df)} rows, {len(df.columns)} columns")

# ── Column aliases (handle whatever names are present) ────────────
def _col(df, *names):
    for n in names:
        if n in df.columns:
            return n
    return None

COL_ORIG    = _col(df, "translated discharge_summary", "original", "discharge_summary")
COL_RULE    = _col(df, "rule_based_simplified", "lay_replaced_summary")
COL_FINAL   = _col(df, "indian_lay_english", "simplified", "lay_english")
COL_FLESCH  = _col(df, "flesch_simplified", "flesch_improvement")
COL_RATIO   = _col(df, "compression_ratio")
COL_DIAGN   = _col(df, "translated diagnosis_ICD", "diagnosis_ICD")
COL_OUTCOME = _col(df, "translanted outcome", "outcome")

# ── Detect regional language columns ──────────────────────────────
LANG_COLS = {c: c.replace("lang_", "").title()
             for c in df.columns if c.startswith("lang_")}
IS_MULTILINGUAL = len(LANG_COLS) > 0

print(f"  Original column  : {COL_ORIG}")
print(f"  Rule-based column: {COL_RULE}")
print(f"  Final lay column : {COL_FINAL}")
if IS_MULTILINGUAL:
    print(f"  Language columns : {list(LANG_COLS.values())}")

# ── 1. HTML REPORT ────────────────────────────────────────────────
def esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")

if IS_MULTILINGUAL:
    # ── Multilingual HTML with language tabs ──────────────────────
    lang_display = {"indian_lay_english": "English (Lay)"}
    lang_display.update(LANG_COLS)

    tab_buttons = []
    tab_bodies  = []

    for tab_idx, (col_key, display_name) in enumerate(lang_display.items()):
        active_cls = "active" if tab_idx == 0 else ""
        tab_buttons.append(
            f'<button class="tab-btn {active_cls}" '
            f'onclick="showTab(\'{col_key}\')" id="btn_{col_key}">'
            f'{display_name}</button>'
        )

        rows_html = []
        for i, row in df.iterrows():
            orig  = esc(row[COL_ORIG]) if COL_ORIG else ""
            trans = esc(row[col_key]) if col_key in df.columns else "[N/A]"
            bg    = "#f4f8fe" if i % 2 == 0 else "#ffffff"
            rows_html.append(f"""
    <tr style='background:{bg}'>
      <td class='num'>{i+1}</td>
      <td class='text orig'>{orig}</td>
      <td class='text translated'>{trans}</td>
    </tr>""")

        display = "block" if tab_idx == 0 else "none"
        tab_bodies.append(f"""
    <div class="tab-content" id="tab_{col_key}" style="display:{display}">
      <table>
      <thead><tr>
        <th>#</th>
        <th>Original Discharge Summary</th>
        <th style='background:#145a32'>{display_name}</th>
      </tr></thead>
      <tbody>{''.join(rows_html)}</tbody>
      </table>
    </div>""")

    lang_list = ', '.join(lang_display.values())
    html = f"""<!DOCTYPE html>
<html lang='en'>
<head>
<meta charset='UTF-8'>
<title>Multilingual Medical Pipeline Results</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13.5px; line-height: 1.6;
    margin: 0; padding: 20px; background: #eef2f7;
    color: #222;
  }}
  h1 {{ color: #1a3a5c; margin-bottom: 4px; }}
  .subtitle {{ color: #666; margin-bottom: 16px; font-size: 13px; }}
  .search-bar {{
    margin-bottom: 14px;
    display: flex; gap: 10px; align-items: center;
  }}
  .search-bar input {{
    padding: 7px 12px; border: 1px solid #bdd0e8;
    border-radius: 6px; font-size: 13px; width: 320px;
    outline: none;
  }}
  .search-bar input:focus {{ border-color: #1a3a5c; }}
  .tabs {{ display: flex; gap: 4px; margin: 16px 0 0 0; flex-wrap: wrap; }}
  .tab-btn {{ padding: 8px 18px; border: none; border-radius: 6px 6px 0 0;
              cursor: pointer; font-size: 13px; font-weight: 600;
              background: #d6e4f0; color: #1a3a5c; transition: .15s; }}
  .tab-btn.active {{ background: #1a3a5c; color: #fff; }}
  .tab-btn:hover  {{ background: #a8c4df; }}
  table {{
    border-collapse: collapse; width: 100%; background: #fff;
    box-shadow: 0 2px 8px rgba(0,0,0,.1); border-radius: 0 8px 8px 8px;
    overflow: hidden;
  }}
  thead th {{
    background: #1a3a5c; color: #fff; padding: 11px 14px;
    text-align: left; position: sticky; top: 0; z-index: 3;
    white-space: nowrap; font-size: 13px;
  }}
  td {{ padding: 11px 14px; vertical-align: top;
        border-bottom: 1px solid #dde8f5; }}
  .num  {{ width: 42px; text-align: center; color: #999;
           white-space: nowrap; font-size: 12px; }}
  .text {{ max-width: 450px; line-height: 1.55; }}
  .orig {{ color: #333; }}
  .translated {{ color: #145a32; font-weight: 600; }}
  tr:hover td {{ filter: brightness(0.96); }}
  .hidden {{ display: none; }}
</style>
<script>
function showTab(langKey) {{
  document.querySelectorAll('.tab-content').forEach(function(el) {{ el.style.display='none'; }});
  document.querySelectorAll('.tab-btn').forEach(function(el) {{ el.classList.remove('active'); }});
  document.getElementById('tab_'+langKey).style.display='block';
  document.getElementById('btn_'+langKey).classList.add('active');
}}
function filterTable() {{
  var q = document.getElementById('search').value.toLowerCase();
  document.querySelectorAll('tbody tr').forEach(function(row) {{
    row.classList.toggle('hidden', q && !row.innerText.toLowerCase().includes(q));
  }});
}}
</script>
</head>
<body>
<h1>Medical Discharge Summary — Multilingual Translation</h1>
<div class='subtitle'>{len(df)} summaries &times; {len(lang_display)} languages ({lang_list})</div>
<div class='search-bar'>
  <input id='search' type='text' placeholder='Search any row...' oninput='filterTable()'>
</div>
<div class="tabs">{''.join(tab_buttons)}</div>
{''.join(tab_bodies)}
</body></html>"""

else:
    # ── English-only HTML (no language columns) ───────────────────
    rows_html = []
    for i, row in df.iterrows():
        orig    = esc(row[COL_ORIG])   if COL_ORIG   else ""
        rule_b  = esc(row[COL_RULE])   if COL_RULE   else ""
        final   = esc(row[COL_FINAL])  if COL_FINAL  else ""
        diagn   = esc(row[COL_DIAGN])  if COL_DIAGN  else ""
        outcome = esc(row[COL_OUTCOME]) if COL_OUTCOME else ""
        flesch  = f"{row[COL_FLESCH]:.1f}" if COL_FLESCH and pd.notna(row[COL_FLESCH]) else ""
        ratio   = f"{row[COL_RATIO]:.2f}"  if COL_RATIO  and pd.notna(row[COL_RATIO])  else ""
        bg      = "#f4f8fe" if i % 2 == 0 else "#ffffff"

        rows_html.append(f"""
    <tr>
      <td class='num' style='background:{bg}'>{i+1}</td>
      <td class='text orig' style='background:{bg}'>{orig}</td>
      {'<td class="text simp" style="background:' + bg + '">' + rule_b + '</td>' if rule_b else ''}
      <td class='text final' style='background:{bg}'>{final}</td>
      <td class='meta' style='background:{bg}'>{diagn}</td>
      <td class='meta' style='background:{bg}'>{outcome}</td>
      <td class='num'  style='background:{bg}'>{flesch}</td>
      <td class='num'  style='background:{bg}'>{ratio}</td>
    </tr>""")

    simp_header = "<th>Rule-Based Simplified</th>" if COL_RULE else ""

    html = f"""<!DOCTYPE html>
<html lang='en'>
<head>
<meta charset='UTF-8'>
<title>Medical Pipeline Results</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13.5px; line-height: 1.6;
    margin: 0; padding: 20px; background: #eef2f7;
    color: #222;
  }}
  h1 {{ color: #1a3a5c; margin-bottom: 4px; }}
  .subtitle {{ color: #666; margin-bottom: 16px; font-size: 13px; }}
  .search-bar {{
    margin-bottom: 14px;
    display: flex; gap: 10px; align-items: center;
  }}
  .search-bar input {{
    padding: 7px 12px; border: 1px solid #bdd0e8;
    border-radius: 6px; font-size: 13px; width: 320px;
    outline: none;
  }}
  .search-bar input:focus {{ border-color: #1a3a5c; }}
  .legend {{ font-size: 12px; color: #555; }}
  .legend span {{ display: inline-block; padding: 2px 8px;
    border-radius: 4px; margin-right: 6px; }}
  table {{
    border-collapse: collapse; width: 100%; background: #fff;
    box-shadow: 0 2px 8px rgba(0,0,0,.1); border-radius: 8px;
    overflow: hidden;
  }}
  thead th {{
    background: #1a3a5c; color: #fff; padding: 11px 14px;
    text-align: left; position: sticky; top: 0; z-index: 3;
    white-space: nowrap; font-size: 13px;
  }}
  td {{ padding: 11px 14px; vertical-align: top;
        border-bottom: 1px solid #dde8f5; }}
  .num  {{ width: 42px; text-align: center; color: #999;
           white-space: nowrap; font-size: 12px; }}
  .text {{ max-width: 380px; }}
  .meta {{ max-width: 160px; font-size: 12px; color: #555; }}
  .orig  {{ color: #333; }}
  .simp  {{ color: #1a4d7a; background-color: inherit !important; }}
  .final {{ color: #145a32; font-weight: 600;
            background-color: #f0fff4 !important; }}
  tr:hover td {{ filter: brightness(0.96); }}
  .hidden {{ display: none; }}
</style>
<script>
function filterTable() {{
  var q = document.getElementById('search').value.toLowerCase();
  var rows = document.querySelectorAll('tbody tr');
  rows.forEach(function(row) {{
    row.classList.toggle('hidden', q && !row.innerText.toLowerCase().includes(q));
  }});
}}
</script>
</head>
<body>
<h1>Medical Discharge Summary → Indian Lay English</h1>
<div class='subtitle'>{len(df)} summaries &nbsp;|&nbsp; Source: {os.path.basename(INPUT_PATH)}</div>
<div class='search-bar'>
  <input id='search' type='text' placeholder='Search any row...' oninput='filterTable()'>
  <span class='legend'>
    <span style='background:#f0fff4;color:#145a32;border:1px solid #a0d8b3'>■ Final Lay English</span>
    <span style='background:#eef5fb;color:#1a4d7a;border:1px solid #aac4e0'>■ Rule-Based</span>
  </span>
</div>
<table>
<thead>
<tr>
  <th>#</th>
  <th>Original Summary</th>
  {simp_header}
  <th>Final Indian Lay English</th>
  <th>Diagnosis</th>
  <th>Outcome</th>
  <th>Flesch</th>
  <th>Ratio</th>
</tr>
</thead>
<tbody>
{''.join(rows_html)}
</tbody>
</table>
</body>
</html>"""

with open(HTML_OUT, "w", encoding="utf-8") as f:
    f.write(html)
print(f"\n  OK  HTML report saved  -> {HTML_OUT}")

# ── 2. FORMATTED EXCEL ────────────────────────────────────────────
LONG_COLS   = {COL_ORIG, COL_RULE, COL_FINAL, "medical_entities", "expanded_summary"} - {None}
LONG_COLS  |= set(LANG_COLS.keys())   # regional language columns are long text too
NARROW_COLS = {"flesch_original", "flesch_simplified", "flesch_improvement",
               "flesch_rule_based", "word_count_original", "word_count_simplified",
               "compression_ratio", "lenght of stay"}

df.to_excel(XLSX_OUT, index=False, engine="openpyxl")

wb = openpyxl.load_workbook(XLSX_OUT)
ws = wb.active

HDR_FILL   = PatternFill("solid", fgColor="1A3A5C")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
ODD_FILL   = PatternFill("solid", fgColor="F4F8FE")
EVN_FILL   = PatternFill("solid", fgColor="FFFFFF")
FINAL_FILL = PatternFill("solid", fgColor="F0FFF4")
WRAP       = Alignment(wrap_text=True, vertical="top")
TOP        = Alignment(wrap_text=False, vertical="top")
CENTER_HDR = Alignment(wrap_text=False, vertical="center", horizontal="center")

col_names = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

# Header row
for ci, name in enumerate(col_names, 1):
    cell = ws.cell(1, ci)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER_HDR
    letter = get_column_letter(ci)
    if name in LONG_COLS:
        ws.column_dimensions[letter].width = 70
    elif name in NARROW_COLS:
        ws.column_dimensions[letter].width = 15
    else:
        ws.column_dimensions[letter].width = 22

# Data rows
for ri in range(2, ws.max_row + 1):
    base_fill = ODD_FILL if ri % 2 == 0 else EVN_FILL
    for ci, name in enumerate(col_names, 1):
        cell = ws.cell(ri, ci)
        if name == COL_FINAL:
            cell.fill = FINAL_FILL
        else:
            cell.fill = base_fill
        cell.alignment = WRAP if name in LONG_COLS else TOP
    ws.row_dimensions[ri].height = 120  # ~8 lines visible

ws.freeze_panes = "A2"
wb.save(XLSX_OUT)
print(f"  OK  Formatted Excel saved -> {XLSX_OUT}")
print()
print("  TIP: For best readability, open results_report.html in Chrome/Edge.")
